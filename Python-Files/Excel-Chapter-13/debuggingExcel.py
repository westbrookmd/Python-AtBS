import logging
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s -  %(levelname)s -  %(message)s')
#logging.disable(logging.CRITICAL)

logging.debug('Start of program')

# Requires import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
# Print worksheet names
logging.debug("List of Sheet Titles = " + str(wb.sheetnames))
# Get a sheet
sheet = wb['Sheet1']
logging.debug("Sheet = " + str(sheet))
# Get a cell from a sheet
cell = sheet['A1']
logging.debug("Cell = " + str(cell))
# Get the cell value
logging.debug("Cell Value = " + str(cell.value))

#each % inserts the value from the corresponding values from the second set of parameters
logging.debug('Row %s, Column %s is %s' % (cell.row, cell.column, cell.value))

logging.debug("Sheet Highest Column Number = " + str(sheet.max_column))
logging.debug("Sheet Lowest Column Number = " + str(sheet.min_column))
logging.debug("Column Letter 1 = " + str(get_column_letter(1)))

#Requires from openpyxl.utils import get_column_letter, column_index_from_string
logging.debug("Sheet Highest Column Letter = " + str(get_column_letter(sheet.max_column)))
logging.debug("Sheet Minimum Column Letter = " + str(get_column_letter(sheet.min_column)))

logging.debug((tuple(sheet['A1':'C3']))) #Get all cells from A1 to C3
print("Printing all cells from A1 to C3 by Row")
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')
print("Done.\n")

for cellObj in list(sheet.columns)[1]:
        print(cellObj.value)
